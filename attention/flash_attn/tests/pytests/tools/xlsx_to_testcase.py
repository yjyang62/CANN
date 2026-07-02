#!/usr/bin/python
# -*- coding: utf-8 -*-
# Copyright (c) 2025 Huawei Technologies Co., Ltd.
# ======================================================================================================================
"""
将 flash_attn 红线用例 xlsx 表格转换为 test_cases 格式的 Python 文件。

用法:
    python xlsx_to_testcase.py --input <xlsx文件> [--output <py文件>] [--sheet <sheet名>]

生成的 .py 文件包含 TestCases 字典，可直接被 test_flash_attn.py 导入：
    python test_flash_attn.py --case_files <生成的模块名(不含.py)>

示例:
    python xlsx_to_testcase.py -i flash_attn_redline.xlsx
    python xlsx_to_testcase.py -i flash_attn_redline.xlsx -o redline.py
    python xlsx_to_testcase.py -i flash_attn_redline.xlsx -s Sheet2

依赖:
    pip install openpyxl
"""

import argparse
import os
import sys
import re

try:
    import openpyxl
except ImportError:
    print("[ERROR] 需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)


# ======================== xlsx 列名 → test_cases 模块 key 映射 ========================
# None = 跳过该列；以 '_' 开头的内部标记不直接写入输出
COLUMN_MAP = {
    # ---- 格式1: 红线用例表 ----
    'Testcase_Name':       '_key_',
    'custorm_info':        None,
    'batch_size':          'B',
    'num_heads_q':         'N1',
    'num_heads_kv':        'N2',
    'Q_S':                 'S1',
    'KV_S':                'S2',
    'layout_q':            'layout_q',
    'layout_kv':           'layout_kv',
    'layout_out':          'layout_out',
    'head_dim':            'D',
    'q_dtype':             'Dtype',
    'q_datarange':         'q_range',
    'k_datarange':         'k_range',
    'v_datarange':         'v_range',
    'softmax_scale':       'scale',
    'mask_mode':           'mask_mode',
    'win_left':            'winLeft',
    'win_right':           'winRight',
    'seqused_q':           'seqused_q',
    'seqused_kv':          'seqused_kv',
    'cu_seqlens_q':        'cu_seqlens_q',
    'cu_seqlens_kv':       'cu_seqlens_kv',
    'blockSize':           'block_size',
    'block_table_shape':   '_block_table_shape_',
    'block_table_dtype':   None,
    'return_softmax_lse':  'return_softmax_lse',
    'attn_out_shape':      None,
    'attn_out_dtype':      None,
    'softmax_lse':         None,
    'softmaxLse_dtype':    None,
    'q_shape':             None,
    'k_shape':             None,
    'k_dtype':             None,
    'v_shape':             None,
    'v_dtype':             None,
    'attn_mask_shape':     None,
    'attn_mask_dtype':     None,
    'learnable_sink':      None,
    'metadata':            None,
    'max_seqlen_q':        'max_seqlen_q',
    'max_seqlen_kv':       'max_seqlen_kv',
    # ---- 格式2: varlen 测试表 ----
    'testcase_name':       '_key_',
    'enable':              '_enable_',
    'level':               None,
    'B':                   'B',
    'N1':                  'N1',
    'N2':                  'N2',
    'S1':                  'S1',
    'S2':                  'S2',
    'D':                   'D',
    'seqlens_list_q':      '_seqlens_list_q_',   # 非前缀和，需转换
    'seqlens_list_kv':     '_seqlens_list_kv_',  # 非前缀和，需转换
    'dtype':               'Dtype',
    'sparse_mode':         'mask_mode',
    'pre_tokens':          'win_left',
    'next_tokens':         'win_right',
    'input_layout':        '_input_layout_',      # 同时设置 layout_q/kv/out
    'atten_mask_dtype':    'atten_mask_dtype',
    'atten_mask_layout':   'atten_mask_layout',
}

DTYPE_MAP = {
    'BF16': 'bf16', 'FP16': 'fp16', 'FP32': 'fp32', 'INT8': 'int8',
    'bf16': 'bf16', 'fp16': 'fp16', 'fp32': 'fp32', 'int8': 'int8',
    'FLOAT16': 'fp16', 'BFLOAT16': 'bf16',
    'float16': 'fp16', 'bfloat16': 'bf16',
}

# 输出时 key 的排列顺序（与 test_cases 模块风格一致）
KEY_ORDER = [
    'B', 'N1', 'N2', 'S1', 'S2', 'D',
    'layout_q', 'layout_kv', 'layout_out',
    'Dtype',
    'q_range', 'k_range', 'v_range',
    'scale',
    'mask_mode', 'win_left', 'win_right',
    'winLeft', 'winRight',
    'seqused_q', 'seqused_kv',
    'max_seqlen_q', 'max_seqlen_kv',
    'cu_seqlens_q', 'cu_seqlens_kv',
    'block_table', 'block_size',
    'atten_mask_dtype', 'atten_mask_layout',
    'return_softmax_lse',
]


# ======================== 辅助函数 ========================

def is_empty(val):
    """判断 xlsx 单元格值是否为空 / 不适用"""
    if val is None:
        return True
    s = str(val).strip()
    return s in ('', '/', 'None', 'none', 'N/A', 'n/a', '-')


def try_int(v):
    """能转 int 就转 int，否则返回 float"""
    f = float(v)
    if f == int(f) and abs(f) < 1e15:
        return int(f)
    return f


def parse_number_list(val_str):
    """'0,128,256' → [0, 128, 256]"""
    parts = [x.strip() for x in str(val_str).split(',') if x.strip()]
    return [try_int(x) for x in parts]


def parse_int_list(val_str):
    """'0,128,256' → [0, 128, 256] (强制 int)"""
    parts = [x.strip() for x in str(val_str).split(',') if x.strip()]
    return [int(float(x)) for x in parts]


def parse_data_range(val):
    """'-10,10' → (-10.0, 10.0)"""
    s = str(val).strip()
    parts = [x.strip() for x in s.split(',') if x.strip()]
    if len(parts) == 2:
        return (float(parts[0]), float(parts[1]))
    return None


def generate_block_table(shape_str, batch_size):
    """从 block_table_shape '4,8' 生成 [[0..7],[8..15],...]"""
    parts = [int(x.strip()) for x in shape_str.split(',') if x.strip()]
    if len(parts) != 2:
        return None
    b, num_blocks = parts
    table = []
    idx = 0
    for _ in range(b):
        table.append(list(range(idx, idx + num_blocks)))
        idx += num_blocks
    return table


def parse_bracket_list(val):
    """解析 '[54, 341, 57, 46]' 或 '54, 341, 57, 46' 格式为 int 列表"""
    s = str(val).strip()
    # 去除外层方括号
    s = s.strip('[]')
    parts = [x.strip() for x in s.split(',') if x.strip()]
    return [int(float(x)) for x in parts]


def seqlens_to_prefix_sum(lengths):
    """将逐 batch 长度列表转换为前缀和（含首位 0）
    [54, 341, 57] → [0, 54, 395, 452]
    """
    result = [0]
    for l in lengths:
        result.append(result[-1] + l)
    return result


# ======================== 行转换 ========================

def convert_row(headers, row_values):
    """将 xlsx 一行数据转换为 (case_name, case_dict)"""
    raw = {}
    for i, h in enumerate(headers):
        if i < len(row_values):
            raw[h] = row_values[i]

    # 提取 case_name（兼容两种列名）
    case_name = None
    for key_col in ('Testcase_Name', 'testcase_name'):
        if key_col in raw and not is_empty(raw[key_col]):
            case_name = str(raw[key_col]).strip()
            break

    # enable 列过滤（格式2）
    if 'enable' in raw and not is_empty(raw['enable']):
        enable_str = str(raw['enable']).strip().lower()
        if enable_str in ('false', '0', 'no'):
            return None, None

    batch_size = 1
    bt_shape = None

    if 'batch_size' in raw and not is_empty(raw['batch_size']):
        batch_size = int(float(raw['batch_size']))
    if 'B' in raw and not is_empty(raw['B']):
        batch_size = int(float(raw['B']))
    if 'block_table_shape' in raw and not is_empty(raw['block_table_shape']):
        bt_shape = str(raw['block_table_shape']).strip()

    if case_name is None:
        return None, None

    case_dict = {}

    # ---- 处理 input_layout → layout_q/kv/out（格式2）----
    if 'input_layout' in raw and not is_empty(raw['input_layout']):
        layout_val = str(raw['input_layout']).strip()
        case_dict['layout_q'] = [layout_val]
        case_dict['layout_kv'] = [layout_val]
        case_dict['layout_out'] = [layout_val]

    # ---- 处理 seqlens_list_q/kv → cu_seqlens_q/kv（逐batch长度→前缀和）----
    if 'seqlens_list_q' in raw and not is_empty(raw['seqlens_list_q']):
        lengths_q = parse_bracket_list(raw['seqlens_list_q'])
        case_dict['cu_seqlens_q'] = [seqlens_to_prefix_sum(lengths_q)]
    if 'seqlens_list_kv' in raw and not is_empty(raw['seqlens_list_kv']):
        lengths_kv = parse_bracket_list(raw['seqlens_list_kv'])
        case_dict['cu_seqlens_kv'] = [seqlens_to_prefix_sum(lengths_kv)]

    for col_name, val in raw.items():
        if col_name not in COLUMN_MAP:
            continue
        key = COLUMN_MAP[col_name]
        if key is None or (isinstance(key, str) and key.startswith('_')):
            continue
        if is_empty(val):
            continue

        val_str = str(val).strip()

        # 跳过已由上面特殊逻辑处理的字段
        if key in ('layout_q', 'layout_kv', 'layout_out') and key in case_dict:
            continue
        if key in ('cu_seqlens_q', 'cu_seqlens_kv') and key in case_dict:
            continue

        # ---- 整数标量 ----
        if key in ('B', 'N1', 'N2', 'S1', 'S2', 'D', 'mask_mode', 'block_size',
                   'win_left', 'win_right', 'max_seqlen_q', 'max_seqlen_kv'):
            case_dict[key] = [int(float(val))]

        # ---- 字符串标量 (layout) ----
        elif key in ('layout_q', 'layout_kv', 'layout_out'):
            case_dict[key] = [val_str]

        # ---- dtype ----
        elif key == 'Dtype':
            case_dict[key] = [DTYPE_MAP.get(val_str.upper(), val_str.lower())]

        # ---- 数据范围 (-10,10) → (-10.0, 10.0) ----
        elif key in ('q_range', 'k_range', 'v_range'):
            r = parse_data_range(val)
            if r is not None:
                case_dict[key] = [r]

        # ---- 浮点标量 (softmax_scale) ----
        elif key in ('scale',):
            case_dict[key] = [float(val)]

        # ---- winLeft / winRight ----
        elif key in ('winLeft', 'winRight'):
            case_dict[key] = [int(float(val))]

        # ---- seqused_q / seqused_kv → [int, ...] ----
        elif key in ('seqused_q', 'seqused_kv'):
            case_dict[key] = [parse_int_list(val_str)]

        # ---- cu_seqlens_q / cu_seqlens_kv → [int, ...] ----
        elif key in ('cu_seqlens_q', 'cu_seqlens_kv'):
            case_dict[key] = [parse_int_list(val_str)]

        # ---- 布尔 ----
        elif key in ('return_softmax_lse',):
            if val_str.lower() in ('true', '1', 'yes'):
                case_dict[key] = [True]
            elif val_str.lower() in ('false', '0', 'no'):
                case_dict[key] = [False]

        # ---- 字符串透传 (atten_mask_dtype, atten_mask_layout 等) ----
        elif key in ('atten_mask_dtype', 'atten_mask_layout'):
            case_dict[key] = [val_str]

    # block_table 由 shape 自动生成
    if bt_shape:
        bt = generate_block_table(bt_shape, batch_size)
        if bt:
            case_dict['block_table'] = [bt]

    return case_name, case_dict


# ======================== 代码格式化 ========================

def fmt_val(v):
    """将 Python 值格式化为可直接写入 .py 的字符串"""
    if isinstance(v, bool):
        return 'True' if v else 'False'
    if isinstance(v, str):
        return repr(v)
    if isinstance(v, tuple):
        parts = []
        for x in v:
            if isinstance(x, float) and x == int(x):
                parts.append(str(int(x)))
            else:
                parts.append(repr(x))
        return f"({', '.join(parts)})"
    if isinstance(v, list):
        inner = ', '.join(fmt_val(x) for x in v)
        return f"[{inner}]"
    if isinstance(v, float):
        if v == int(v) and abs(v) < 1e15:
            return str(int(v))
        return repr(v)
    return repr(v)


def format_case(case_dict):
    """生成一个 case 的缩进代码行"""
    lines = []
    written = set()
    for k in KEY_ORDER:
        if k in case_dict:
            lines.append(f"        '{k}': [{fmt_val(case_dict[k][0])}],")
            written.add(k)
    for k, v in case_dict.items():
        if k not in written:
            lines.append(f"        '{k}': [{fmt_val(v[0])}],")
    return '\n'.join(lines)


# ======================== 主流程 ========================

def xlsx_to_testcase(input_xlsx, output_py, sheet_name=None):
    wb = openpyxl.load_workbook(input_xlsx, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    print(f"[INFO] 读取 sheet: '{ws.title}'")

    # 读取表头（第 1 行）
    headers = [str(c.value).strip() if c.value else '' for c in ws[1]]

    # 提示未识别的列名
    unknown = [h for h in headers if h and h not in COLUMN_MAP]
    if unknown:
        print(f"[INFO] 以下列名未在映射中，将被忽略: {unknown}")

    has_key_col = any(h in ('Testcase_Name', 'testcase_name') for h in headers)
    if not has_key_col:
        print(f"[ERROR] 缺少必需列: Testcase_Name 或 testcase_name")
        wb.close()
        sys.exit(1)

    # 逐行转换
    cases = {}
    skip_count = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_values = list(row)
        if all(is_empty(v) for v in row_values):
            continue

        name, cdict = convert_row(headers, row_values)
        if name is None:
            print(f"[WARN] 第 {row_idx} 行缺少 Testcase_Name，跳过")
            skip_count += 1
            continue
        if not cdict:
            print(f"[WARN] 第 {row_idx} 行 '{name}' 无有效参数，跳过")
            skip_count += 1
            continue
        if name in cases:
            print(f"[WARN] 用例名重复: '{name}'（第 {row_idx} 行），将覆盖前一条")
        cases[name] = cdict

    wb.close()

    if not cases:
        print("[ERROR] 未解析到任何有效用例")
        sys.exit(1)

    # 写 .py 文件
    with open(output_py, 'w', encoding='utf-8') as f:
        f.write('#!/usr/bin/python\n')
        f.write('# -*- coding: utf-8 -*-\n')
        f.write(f'# Auto-generated from {os.path.basename(input_xlsx)} by xlsx_to_testcase.py\n')
        f.write('# DO NOT EDIT — regenerate from source xlsx instead.\n')
        f.write('#\n')
        f.write(f'# Total cases: {len(cases)}\n')
        f.write('\n')
        f.write('TestCases = {\n')
        for name, cdict in cases.items():
            # 用例名中的单引号转义
            safe_name = name.replace("'", "\\'")
            f.write(f"    '{safe_name}': {{\n")
            f.write(format_case(cdict))
            f.write(f"\n    }},\n")
        f.write('}\n')

    print(f"[OK] 已生成 {len(cases)} 个用例 → {output_py}")
    if skip_count:
        print(f"[INFO] 跳过 {skip_count} 行（空行或无效行）")
    mod = os.path.splitext(os.path.basename(output_py))[0]
    print(f"\n[用法] 运行全部用例:")
    print(f"    python test_flash_attn.py --case_files {mod}")
    print(f"[用法] 运行单个用例:")
    print(f"    python test_flash_attn.py --case_files {mod} --case_id <用例名>")


def main():
    p = argparse.ArgumentParser(
        description='将 flash_attn 红线用例 xlsx 表格转换为 test_cases 格式',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python xlsx_to_testcase.py -i flash_attn_redline.xlsx
  python xlsx_to_testcase.py -i flash_attn_redline.xlsx -o redline.py
  python xlsx_to_testcase.py -i flash_attn_redline.xlsx -s Sheet2
""")
    p.add_argument('--input',  '-i', required=True,  help='输入 xlsx 文件路径')
    p.add_argument('--output', '-o', default=None,
                   help='输出 py 文件路径 (默认: 同目录 <xlsx名>.py)')
    p.add_argument('--sheet',  '-s', default=None,
                   help='xlsx sheet 名称 (默认: 活动 sheet)')
    args = p.parse_args()

    if not os.path.exists(args.input):
        print(f"[ERROR] 文件不存在: {args.input}")
        sys.exit(1)

    output = args.output
    if output is None:
        base = os.path.splitext(os.path.basename(args.input))[0]
        safe_base = re.sub(r'[^\w]', '_', base)
        output = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              f"{safe_base}.py")

    xlsx_to_testcase(args.input, output, args.sheet)


if __name__ == '__main__':
    main()
